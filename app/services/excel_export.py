"""Exportação do histórico para Excel formatado.

A função `exportar_excel` é chamada automaticamente sempre que se grava
o histórico, mantendo o ficheiro `data/historico_encomendas.xlsx`
sempre sincronizado.
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app import config


# --- Paleta ---
COR_HEADER       = "1A1916"
COR_ACCENT       = "E8B84B"
COR_LINHA_PAR    = "F7F4ED"
COR_LINHA_IMPAR  = "FFFFFF"
COR_AM           = "C9943A"
COR_FR           = "2E8B57"
COR_OUTRO        = "808080"

# Cores por tipo de cliente — fácil de adicionar novos no futuro
COR_POR_TIPO = {
    "AM":    COR_AM,
    "FR":    COR_FR,
    "OUTRO": COR_OUTRO,
}

# --- Estrutura ---
HEADERS    = ["Nº", "Data", "Hora", "Cliente", "Tipo", "Produtos", "Total Unid."]
COL_WIDTHS = [6,    12,     8,      28,        7,      55,         13]


def exportar_excel(historico: list) -> None:
    """Gera o ficheiro Excel formatado com o histórico completo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico de Encomendas"

    _escrever_titulo(ws)
    _escrever_cabecalhos(ws)
    _escrever_dados(ws, historico)
    if historico:
        _escrever_total(ws, historico)

    ws.freeze_panes = "A5"
    wb.save(str(config.EXCEL_EXPORT))


# --- Secções privadas ---

def _escrever_titulo(ws):
    ws.merge_cells("A1:G1")
    ws["A1"] = "Casa do Apicultor — Histórico de Encomendas"
    ws["A1"].font = Font(name="Arial", bold=True, color="1A1916", size=14)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].fill = PatternFill("solid", start_color="FDF8EE")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Exportado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    ws["A2"].font = Font(name="Arial", size=9, color="999999", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6


def _escrever_cabecalhos(ws):
    border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
    )
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, color=COR_ACCENT, size=11)
        cell.fill = PatternFill("solid", start_color=COR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 22


def _escrever_dados(ws, historico):
    border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
    )
    normal_font = Font(name="Arial", size=10)
    bold_font   = Font(name="Arial", bold=True, size=10)

    for row_idx, enc in enumerate(historico, start=5):
        cor = COR_LINHA_PAR if row_idx % 2 == 0 else COR_LINHA_IMPAR
        fill = PatternFill("solid", start_color=cor)

        produtos_txt = "\n".join(
            f"• {p['nome']}  ×{p['quantidade']}"
            + (f"  — {p['notas']}" if p.get("notas") else "")
            for p in enc["produtos"]
        )
        total = sum(p["quantidade"] for p in enc["produtos"])

        valores = [
            f"{enc['id']:04d}", enc["data"], enc["hora"], enc["cliente"],
            enc["cliente_tipo"], produtos_txt, total,
        ]

        ws.row_dimensions[row_idx].height = max(18, len(enc["produtos"]) * 16)

        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if col_idx in (1, 3, 5, 7) else "left",
                vertical="center",
                wrap_text=True,
            )
            if col_idx == 5:  # Tipo (AM/FR/OUTRO)
                cell.font = Font(
                    name="Arial", bold=True, size=10,
                    color=COR_POR_TIPO.get(valor, COR_HEADER),
                )
            elif col_idx == 7:
                cell.font = bold_font
            else:
                cell.font = normal_font


def _escrever_total(ws, historico):
    total_row = len(historico) + 5
    ws.merge_cells(f"A{total_row}:F{total_row}")

    label = ws[f"A{total_row}"]
    label.value = f"Total de encomendas: {len(historico)}"
    label.font = Font(name="Arial", bold=True, size=10, color=COR_HEADER)
    label.fill = PatternFill("solid", start_color="FDF8EE")
    label.alignment = Alignment(horizontal="right", vertical="center")

    cell_total = ws.cell(
        row=total_row, column=7,
        value=f"=SUM(G5:G{total_row - 1})",
    )
    cell_total.font = Font(name="Arial", bold=True, size=10, color=COR_HEADER)
    cell_total.fill = PatternFill("solid", start_color="FDF8EE")
    cell_total.alignment = Alignment(horizontal="center", vertical="center")
    cell_total.border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
    )
    ws.row_dimensions[total_row].height = 20