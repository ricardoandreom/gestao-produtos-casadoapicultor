from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from typing import List, Optional
import json
import os
from datetime import datetime
from pathlib import Path
from docx_generator import gerar_documento_word, gerar_etiquetas, gerar_acabamento
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

app = FastAPI(title="Gestão de Encomendas - Colmeias")

BASE_DIR = Path(__file__).parent
DOCS_DIR = BASE_DIR / "encomendas_docs"
DOCS_DIR_MAQUINACAO = DOCS_DIR / "maquinacao"
DOCS_DIR_ETIQUETAS  = DOCS_DIR / "etiquetas"
DOCS_DIR_ACABAMENTO = DOCS_DIR / "acabamentos"
DATA_FILE = BASE_DIR / "data" / "historico.json"
PRODUCTS_FILE = BASE_DIR / "data" / "products_mapping.xlsx"

DOCS_DIR.mkdir(exist_ok=True)
DOCS_DIR_MAQUINACAO.mkdir(exist_ok=True)
DOCS_DIR_ETIQUETAS.mkdir(exist_ok=True)
DOCS_DIR_ACABAMENTO.mkdir(exist_ok=True)
DATA_FILE.parent.mkdir(exist_ok=True)

if not DATA_FILE.exists():
    DATA_FILE.write_text(json.dumps([]))

app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


# --- Modelos ---

class ProdutoItem(BaseModel):
    nome: str
    quantidade: int
    notas: str = ""

class EncomendaRequest(BaseModel):
    cliente: str
    cliente_tipo: str  # "AM" ou "FR"
    produtos: List[ProdutoItem]


# --- Helpers ---

def carregar_historico():
    return json.loads(DATA_FILE.read_text())

def guardar_historico(historico):
    DATA_FILE.write_text(json.dumps(historico, ensure_ascii=False, indent=2))
    exportar_excel(historico)


def exportar_excel(historico: list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico de Encomendas"

    # Estilos
    cor_header     = "1A1916"
    cor_accent     = "E8B84B"
    cor_linha_par  = "F7F4ED"
    cor_linha_impar = "FFFFFF"

    header_font   = Font(name="Arial", bold=True, color=cor_accent, size=11)
    title_font    = Font(name="Arial", bold=True, color="1A1916", size=14)
    normal_font   = Font(name="Arial", size=10)
    bold_font     = Font(name="Arial", bold=True, size=10)
    border_thin   = Border(
        bottom=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD")
    )

    # Título
    ws.merge_cells("A1:G1")
    ws["A1"] = "Casa do Apicultor — Histórico de Encomendas"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].fill = PatternFill("solid", start_color="FDF8EE")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Exportado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    ws["A2"].font = Font(name="Arial", size=9, color="999999", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # Linha em branco
    ws.row_dimensions[3].height = 6

    # Cabeçalhos
    headers = ["Nº", "Data", "Hora", "Cliente", "Tipo", "Produtos", "Total Unid."]
    col_widths = [6, 12, 8, 28, 7, 55, 13]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = PatternFill("solid", start_color=cor_header)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[4].height = 22

    # Dados
    for row_idx, enc in enumerate(historico, start=5):
        cor_fundo = cor_linha_par if row_idx % 2 == 0 else cor_linha_impar
        fill = PatternFill("solid", start_color=cor_fundo)

        produtos_texto = "\n".join(
            f"• {p['nome']}  ×{p['quantidade']}" + (f"  — {p['notas']}" if p.get('notas') else "")
            for p in enc["produtos"]
        )
        total_unidades = sum(p["quantidade"] for p in enc["produtos"])

        valores = [
            f"{enc['id']:04d}",
            enc["data"],
            enc["hora"],
            enc["cliente"],
            enc["cliente_tipo"],
            produtos_texto,
            total_unidades,
        ]

        n_linhas = len(enc["produtos"])
        ws.row_dimensions[row_idx].height = max(18, n_linhas * 16)

        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.fill = fill
            cell.border = border_thin
            cell.alignment = Alignment(
                horizontal="center" if col_idx in (1, 3, 5, 7) else "left",
                vertical="center",
                wrap_text=True
            )
            if col_idx == 5:  # Tipo AM/FR
                cell.font = Font(name="Arial", bold=True, size=10,
                                 color="C9943A" if valor == "AM" else "2E8B57")
            elif col_idx == 7:
                cell.font = bold_font
            else:
                cell.font = normal_font

    # Linha de totais
    if historico:
        total_row = len(historico) + 5
        ws.merge_cells(f"A{total_row}:F{total_row}")
        cell_label = ws[f"A{total_row}"]
        cell_label.value = f"Total de encomendas: {len(historico)}"
        cell_label.font = Font(name="Arial", bold=True, size=10, color=cor_header)
        cell_label.fill = PatternFill("solid", start_color="FDF8EE")
        cell_label.alignment = Alignment(horizontal="right", vertical="center")

        cell_total = ws.cell(row=total_row, column=7,
                             value=f"=SUM(G5:G{total_row - 1})")
        cell_total.font = Font(name="Arial", bold=True, size=10, color=cor_header)
        cell_total.fill = PatternFill("solid", start_color="FDF8EE")
        cell_total.alignment = Alignment(horizontal="center", vertical="center")
        cell_total.border = border_thin
        ws.row_dimensions[total_row].height = 20

    # Freeze panes abaixo do cabeçalho
    ws.freeze_panes = "A5"

    excel_path = DATA_FILE.parent / "historico_encomendas.xlsx"
    wb.save(str(excel_path))


# --- Rotas ---

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.get("/graficos", response_class=HTMLResponse)
async def graficos(request: Request):
    return templates.TemplateResponse("graficos.html", {"request": request})

@app.get("/mapeamento", response_class=HTMLResponse)
async def mapeamento(request: Request):
    return templates.TemplateResponse("mapeamento.html", {"request": request})


@app.post("/api/encomendas")
async def criar_encomenda(encomenda: EncomendaRequest):
    if encomenda.cliente_tipo not in ("AM", "FR"):
        raise HTTPException(status_code=400, detail="Tipo de cliente inválido. Use 'AM' ou 'FR'.")
    if not encomenda.produtos:
        raise HTTPException(status_code=400, detail="A encomenda deve ter pelo menos um produto.")

    historico = carregar_historico()

    novo_id = (max((e["id"] for e in historico), default=0)) + 1
    data_atual = datetime.now().strftime("%d/%m/%Y")
    hora_atual = datetime.now().strftime("%H:%M")

    nova_encomenda = {
        "id": novo_id,
        "data": data_atual,
        "hora": hora_atual,
        "cliente": encomenda.cliente,
        "cliente_tipo": encomenda.cliente_tipo,
        "produtos": [{"nome": p.nome, "quantidade": p.quantidade, "notas": p.notas} for p in encomenda.produtos],
        "doc_maquinacao": "",
        "doc_etiquetas": "",
        "doc_acabamento": "",
    }

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    data_fmt = datetime.now().strftime('%d%m%Y')

    ficheiros = {
        "doc_maquinacao": (DOCS_DIR_MAQUINACAO / f"maquinacao_{novo_id:04d}_{data_fmt}.docx", gerar_documento_word),
        "doc_etiquetas":  (DOCS_DIR_ETIQUETAS  / f"etiquetas_{novo_id:04d}_{data_fmt}.docx",  gerar_etiquetas),
        "doc_acabamento": (DOCS_DIR_ACABAMENTO / f"acabamento_{novo_id:04d}_{data_fmt}.docx",  gerar_acabamento),
    }

    try:
        for campo, (caminho, func) in ficheiros.items():
            func(nova_encomenda, str(caminho))
            nova_encomenda[campo] = str(caminho)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar documentos Word: {str(e)}")

    historico.append(nova_encomenda)
    guardar_historico(historico)

    return JSONResponse(content={"sucesso": True, "id": novo_id, "mensagem": f"Encomenda #{novo_id} criada com sucesso!"})


@app.get("/api/encomendas")
async def listar_encomendas():
    historico = carregar_historico()
    return JSONResponse(content=historico)


@app.get("/api/encomendas/{encomenda_id}/download/{tipo}")
async def download_encomenda(encomenda_id: int, tipo: str):
    if tipo not in ("maquinacao", "etiquetas", "acabamento"):
        raise HTTPException(status_code=400, detail="Tipo inválido. Use: maquinacao, etiquetas ou acabamento.")

    historico = carregar_historico()
    encomenda = next((e for e in historico if e["id"] == encomenda_id), None)
    if not encomenda:
        raise HTTPException(status_code=404, detail="Encomenda não encontrada.")

    campo = f"doc_{tipo}"
    caminho = encomenda.get(campo, "")

    # Compatibilidade com encomendas antigas que só têm doc_path
    if not caminho and tipo == "maquinacao":
        caminho = encomenda.get("doc_path", "")

    if not caminho or not Path(caminho).exists():
        raise HTTPException(status_code=404, detail=f"Documento '{tipo}' não encontrado.")

    return FileResponse(
        path=caminho,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        filename=Path(caminho).name
    )


@app.get("/api/exportar-excel")
async def download_excel():
    excel_path = DATA_FILE.parent / "historico_encomendas.xlsx"
    if not excel_path.exists():
        historico = carregar_historico()
        exportar_excel(historico)
    return FileResponse(
        path=str(excel_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="historico_encomendas.xlsx"
    )


# --- Rotas: Mapeamento de Produtos ---

class ProductMappingItem(BaseModel):
    nome_produto: str
    ref_interna: str
    categoria_produto: str
    qt_palete_AM_FR: str
    notas: Optional[str] = ""

def carregar_produtos():
    if not PRODUCTS_FILE.exists():
        return []
    df = pd.read_excel(str(PRODUCTS_FILE), dtype=str).fillna("")
    df.columns = [c.strip() for c in df.columns]
    df.columns = ["nome_produto", "ref_interna", "categoria_produto", "qt_palete_AM_FR", "notas"]
    return df.to_dict(orient="records")

def guardar_produto(novo: dict):
    if PRODUCTS_FILE.exists():
        df = pd.read_excel(str(PRODUCTS_FILE), dtype=str).fillna("")
    else:
        df = pd.DataFrame(columns=["nome_produto", "ref_interna", "categoria_produto", "qt_palete_AM_FR", "Notas "])
    col_notas = next((c for c in df.columns if c.strip().lower() == "notas"), "Notas ")
    nova_linha = {
        "nome_produto": novo["nome_produto"],
        "ref_interna": novo["ref_interna"],
        "categoria_produto": novo["categoria_produto"],
        "qt_palete_AM_FR": novo["qt_palete_AM_FR"],
        col_notas: novo.get("notas", ""),
    }
    df = pd.concat([df, pd.DataFrame([nova_linha])], ignore_index=True)
    df.to_excel(str(PRODUCTS_FILE), index=False)

@app.get("/api/produtos")
async def listar_produtos():
    return JSONResponse(content=carregar_produtos())

@app.post("/api/produtos")
async def adicionar_produto(produto: ProductMappingItem):
    guardar_produto(produto.dict())
    return JSONResponse(content={"sucesso": True, "mensagem": "Produto adicionado com sucesso!"})


if __name__ == "__main__":
    import webbrowser
    import threading
    import uvicorn

    def abrir_browser():
        import time
        time.sleep(1.5)
        webbrowser.open("http://localhost:8000")

    threading.Thread(target=abrir_browser, daemon=True).start()
    uvicorn.run(app, host="0.0.0.0", port=8000)